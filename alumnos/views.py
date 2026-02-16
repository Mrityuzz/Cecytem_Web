from django.shortcuts import render, redirect
import openpyxl

# Ruta absoluta al Excel
EXCEL_PATH = r"C:\Users\macie\OneDrive\Escritorio\TI\11mo Cuatrimestre Estadias\cecytem_proyecto\base_datos\BD_Alumnos.xlsx"

def lista_alumnos(request):
    alumnos = []
    try:
        libro = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        hoja = libro.active

        # Leer filas desde la segunda (si la primera es encabezado)
        for i in range(2, hoja.max_row + 1):
            alumnos.append({
                "id": i,
                "nombre": hoja.cell(row=i, column=1).value,
                "num_control": hoja.cell(row=i, column=2).value,
                "telefono": hoja.cell(row=i, column=3).value,
                "carrera": hoja.cell(row=i, column=4).value,
            })

    except Exception as e:
        return render(request, "alumnos/lista.html", {
            "alumnos": [],
            "error": str(e)
        })

    return render(request, "alumnos/lista.html", {"alumnos": alumnos})


def agregar_alumno(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        num_control = request.POST.get("num_control")
        telefono = request.POST.get("telefono")
        carrera = request.POST.get("carrera")

        try:
            libro = openpyxl.load_workbook(EXCEL_PATH)
            hoja = libro.active
            nueva_fila = hoja.max_row + 1

            hoja.cell(row=nueva_fila, column=1, value=nombre)
            hoja.cell(row=nueva_fila, column=2, value=num_control)
            hoja.cell(row=nueva_fila, column=3, value=telefono)
            hoja.cell(row=nueva_fila, column=4, value=carrera)

            libro.save(EXCEL_PATH)

        except Exception as e:
            return render(request, "alumnos/agregar.html", {
                "error": str(e)
            })

        return redirect("lista_alumnos")

    return render(request, "alumnos/agregar.html")


def editar_alumno(request, fila_id):
    libro = openpyxl.load_workbook(EXCEL_PATH)
    hoja = libro.active

    if request.method == "POST":
        hoja.cell(row=fila_id, column=1, value=request.POST.get("nombre"))
        hoja.cell(row=fila_id, column=2, value=request.POST.get("num_control"))
        hoja.cell(row=fila_id, column=3, value=request.POST.get("telefono"))
        hoja.cell(row=fila_id, column=4, value=request.POST.get("carrera"))
        libro.save(EXCEL_PATH)
        return redirect("lista_alumnos")

    alumno = {
        "id": fila_id,
        "nombre": hoja.cell(row=fila_id, column=1).value,
        "num_control": hoja.cell(row=fila_id, column=2).value,
        "telefono": hoja.cell(row=fila_id, column=3).value,
        "carrera": hoja.cell(row=fila_id, column=4).value,
    }
    return render(request, "alumnos/editar.html", {"alumno": alumno})


def eliminar_alumno(request, fila_id):
    libro = openpyxl.load_workbook(EXCEL_PATH)
    hoja = libro.active
    hoja.delete_rows(fila_id)
    libro.save(EXCEL_PATH)
    return redirect("lista_alumnos")
