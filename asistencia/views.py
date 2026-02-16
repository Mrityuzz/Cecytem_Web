from django.shortcuts import render
from django.http import FileResponse, JsonResponse
import openpyxl
import os

# Ruta absoluta al Excel en tu proyecto_cecytem
EXCEL_PATH = r"C:\Users\macie\OneDrive\Escritorio\TI\11mo Cuatrimestre Estadias\cecytem_proyecto\base_datos\BD_Alumnos.xlsx"

def lista_asistencia(request):
    ruta_excel = EXCEL_PATH
    alumnos = []

    if not os.path.isfile(ruta_excel):
        return render(request, "asistencia/lista.html", {
            "alumnos": [],
            "error": "El archivo BD_Alumnos.xlsx no se encontró en la carpeta base_datos de proyecto_cecytem."
        })

    try:
        libro = openpyxl.load_workbook(ruta_excel, data_only=True)
        hoja = libro.active

        # Filtro dinámico por carrera (ejemplo: ?carrera=Electrónica)
        carrera_filtro = request.GET.get("carrera")

        for i in range(2, hoja.max_row + 1):
            alumno = hoja.cell(row=i, column=1).value
            num_control = hoja.cell(row=i, column=2).value
            telefono = hoja.cell(row=i, column=3).value
            carrera = hoja.cell(row=i, column=4).value
            entrada = hoja.cell(row=i, column=5).value
            salida = hoja.cell(row=i, column=6).value

            registro = {
                "alumno": alumno,
                "num_control": num_control,
                "telefono": telefono,
                "carrera": carrera,
                "entrada": entrada,
                "salida": salida,
            }

            if carrera_filtro:
                if carrera and carrera_filtro.lower() in str(carrera).lower():
                    alumnos.append(registro)
            else:
                alumnos.append(registro)

    except Exception as e:
        return render(request, "asistencia/lista.html", {
            "alumnos": [],
            "error": f"Ocurrió un error al abrir el archivo: {e}"
        })

    return render(request, "asistencia/lista.html", {
        "alumnos": alumnos,
        "carrera_filtro": carrera_filtro
    })


# Vista para descargar el Excel original
def descargar_excel(request):
    ruta_excel = EXCEL_PATH
    return FileResponse(open(ruta_excel, "rb"), as_attachment=True, filename="BD_Alumnos.xlsx")


# Vista JSON para actualización en tiempo real
def lista_asistencia_json(request):
    ruta_excel = EXCEL_PATH
    alumnos = []
    libro = openpyxl.load_workbook(ruta_excel, data_only=True)
    hoja = libro.active

    for i in range(2, hoja.max_row + 1):
        alumnos.append({
            "alumno": hoja.cell(row=i, column=1).value,
            "num_control": hoja.cell(row=i, column=2).value,
            "telefono": hoja.cell(row=i, column=3).value,
            "carrera": hoja.cell(row=i, column=4).value,
            "entrada": hoja.cell(row=i, column=5).value,
            "salida": hoja.cell(row=i, column=6).value,
        })

    return JsonResponse({"alumnos": alumnos})
