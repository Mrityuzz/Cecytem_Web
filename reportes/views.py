from django.shortcuts import render
import openpyxl
from collections import Counter

# Ruta absoluta al archivo en cecytem_proyecto
EXCEL_ALUMNOS = r"C:\Users\macie\OneDrive\Escritorio\TI\11mo Cuatrimestre Estadias\cecytem_proyecto\base_datos\BD_Alumnos.xlsx"

def reporte_general(request):
    # --- Alumnos por carrera ---
    libro = openpyxl.load_workbook(EXCEL_ALUMNOS, data_only=True)
    hoja = libro.active
    carreras = []
    for i in range(2, hoja.max_row + 1):  # desde fila 2 para saltar encabezado
        carrera = hoja.cell(row=i, column=4).value
        if carrera:
            carreras.append(carrera)
    conteo_carreras = Counter(carreras)

    contexto = {
        "conteo_carreras": dict(conteo_carreras),
        "total_alumnos": hoja.max_row - 1,  # quitando encabezado
    }
    return render(request, "reportes/reporte_general.html", contexto)
